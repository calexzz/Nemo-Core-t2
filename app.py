from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
import sqlite3
import io
import openpyxl
from datetime import datetime, timedelta
from database import get_db, hash_password, init_db

app = Flask(__name__)
app.secret_key = 'hackathon_t2_secret_key_2026'

# ─── Константы ────────────────────────────────────────────────────────────────

MIN_REST_HOURS = 9       # минимальный непрерывный отдых между сменами
WEEK_NORM_HOURS = 40     # норма часов в неделю
MIN_PASSWORD_LEN = 8     # минимальная длина пароля

# ─── Вспомогательные функции ──────────────────────────────────────────────────

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    from functools import wraps
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if session.get('role') not in roles:
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated
    return decorator

def get_next_14_days():
    today = datetime.today()
    days = []
    weekdays_ru = ['Понедельник','Вторник','Среда','Четверг','Пятница','Суббота','Воскресенье']
    for i in range(14):
        d = today + timedelta(days=i)
        days.append({
            'value': d.strftime('%Y-%m-%d'),
            'label': f"{d.strftime('%Y-%m-%d')} ({weekdays_ru[d.weekday()]})"
        })
    return days

def shift_duration_hours(start_time, end_time):
    """Длительность смены в часах. Возвращает 0 если нет данных."""
    if not start_time or not end_time or start_time == 'Выходной':
        return 0.0
    try:
        sh, sm = map(int, start_time.split(':'))
        eh, em = map(int, end_time.split(':'))
        start_m = sh * 60 + sm
        end_m = eh * 60 + em
        if end_m <= start_m:
            end_m += 1440  # смена переходит через полночь
        return (end_m - start_m) / 60.0
    except Exception:
        return 0.0

def get_week_start(date_str):
    """Возвращает дату понедельника той недели, в которую входит date_str."""
    d = datetime.strptime(date_str, '%Y-%m-%d')
    return (d - timedelta(days=d.weekday())).strftime('%Y-%m-%d')

def calc_week_hours(user_id, week_start, exclude_id=None):
    """Суммирует часы за неделю (пн–вс) для пользователя."""
    week_end = (datetime.strptime(week_start, '%Y-%m-%d') + timedelta(days=6)).strftime('%Y-%m-%d')
    conn = get_db()
    q = 'SELECT start_time, end_time FROM shifts WHERE user_id=? AND shift_date>=? AND shift_date<=? AND start_time!="Выходной"'
    params = [user_id, week_start, week_end]
    if exclude_id:
        q += ' AND id!=?'
        params.append(exclude_id)
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return sum(shift_duration_hours(r['start_time'], r['end_time']) for r in rows)

def check_consecutive_shifts(user_id, new_date, exclude_id=None):
    """Максимальное число смен подряд при добавлении new_date."""
    conn = get_db()
    q = 'SELECT shift_date FROM shifts WHERE user_id=? AND start_time!="Выходной"'
    params = [user_id]
    if exclude_id:
        q += ' AND id!=?'
        params.append(exclude_id)
    rows = conn.execute(q, params).fetchall()
    conn.close()

    dates = sorted(set([r['shift_date'] for r in rows] + [new_date]))
    max_c = current = 1
    for i in range(1, len(dates)):
        d1 = datetime.strptime(dates[i-1], '%Y-%m-%d')
        d2 = datetime.strptime(dates[i], '%Y-%m-%d')
        current = current + 1 if (d2 - d1).days == 1 else 1
        max_c = max(max_c, current)
    return max_c

def check_rest_hours(user_id, new_date, new_start, new_end, exclude_id=None):
    """
    Проверяет соблюдение минимального отдыха (9 ч) между соседними сменами.
    Возвращает (ok: bool, message: str).
    """
    if new_start == 'Выходной' or not new_start or not new_end:
        return True, ''

    conn = get_db()
    q = 'SELECT shift_date, start_time, end_time FROM shifts WHERE user_id=? AND start_time!="Выходной"'
    params = [user_id]
    if exclude_id:
        q += ' AND id!=?'
        params.append(exclude_id)
    rows = conn.execute(q, params).fetchall()
    conn.close()

    # Конвертируем в минуты от эпохи (дата * 1440 + часы*60 + мин)
    def to_abs_minutes(date_str, time_str):
        d = datetime.strptime(date_str, '%Y-%m-%d')
        h, m = map(int, time_str.split(':'))
        return d.toordinal() * 1440 + h * 60 + m

    new_start_abs = to_abs_minutes(new_date, new_start)
    new_end_abs   = to_abs_minutes(new_date, new_end)
    if new_end_abs <= new_start_abs:
        new_end_abs += 1440

    min_gap = MIN_REST_HOURS * 60  # в минутах

    for r in rows:
        if not r['end_time']:
            continue
        s_abs = to_abs_minutes(r['shift_date'], r['start_time'])
        e_abs = to_abs_minutes(r['shift_date'], r['end_time'])
        if e_abs <= s_abs:
            e_abs += 1440

        # Зазор между концом существующей и началом новой
        gap1 = new_start_abs - e_abs
        # Зазор между концом новой и началом существующей
        gap2 = s_abs - new_end_abs

        if 0 < gap1 < min_gap:
            return False, f'Между сменой {r["shift_date"]} и новой сменой менее {MIN_REST_HOURS} ч отдыха.'
        if 0 < gap2 < min_gap:
            return False, f'Между новой сменой и сменой {r["shift_date"]} менее {MIN_REST_HOURS} ч отдыха.'

    return True, ''

def create_notification(user_id, shift_id, shift_date, message):
    """Создаёт уведомление для администраторов."""
    conn = get_db()
    conn.execute(
        'INSERT INTO notifications (user_id, shift_id, shift_date, message, created_at) VALUES (?,?,?,?,?)',
        (user_id, shift_id, shift_date, message, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    )
    conn.commit()
    conn.close()

def get_unread_notifications_count():
    """Кол-во непрочитанных уведомлений для текущего пользователя-админа."""
    if session.get('role') not in ('admin', 'manager'):
        return 0
    conn = get_db()
    row = conn.execute('SELECT COUNT(*) as cnt FROM notifications WHERE is_read=0').fetchone()
    conn.close()
    return row['cnt']

# ─── Маршруты ─────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('role') in ('admin', 'manager'):
        return redirect(url_for('manager_dashboard'))
    return redirect(url_for('employee_schedule'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        conn = get_db()
        user = conn.execute(
            'SELECT * FROM users WHERE username=? AND password=?',
            (username, hash_password(password))
        ).fetchone()
        conn.close()
        if user:
            session['user_id']   = user['id']
            session['username']  = user['username']
            session['full_name'] = user['full_name']
            session['role']      = user['role']
            session['alliance']  = user['alliance']
            session['team']      = user['team']
            return redirect(url_for('index'))
        else:
            error = 'Неверный логин или пароль'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── Страница сотрудника ──────────────────────────────────────────────────────

@app.route('/schedule')
@login_required
def employee_schedule():
    user_id = session['user_id']
    conn = get_db()
    shifts = [dict(r) for r in conn.execute(
        'SELECT * FROM shifts WHERE user_id=? ORDER BY shift_date', (user_id,)
    ).fetchall()]

    # Считаем часы текущей недели для отображения предупреждения
    today = datetime.today().strftime('%Y-%m-%d')
    week_start = get_week_start(today)
    week_hours = calc_week_hours(user_id, week_start)

    # Запросы на переработку сотрудника
    overtime = [dict(r) for r in conn.execute(
        'SELECT * FROM overtime_requests WHERE user_id=? ORDER BY created_at DESC', (user_id,)
    ).fetchall()]
    conn.close()

    return render_template('employee.html',
        shifts=shifts,
        dates=get_next_14_days(),
        user=session,
        week_hours=round(week_hours, 1),
        week_norm=WEEK_NORM_HOURS,
        overtime=overtime
    )

@app.route('/shift/add', methods=['POST'])
@login_required
def add_shift():
    user_id    = session['user_id']
    shift_date = request.form['shift_date']
    start_time = request.form['start_time']
    end_time   = request.form.get('end_time', '')

    conn = get_db()

    # Нет ли уже смены на эту дату
    if conn.execute('SELECT id FROM shifts WHERE user_id=? AND shift_date=?', (user_id, shift_date)).fetchone():
        conn.close()
        return jsonify({'error': 'На эту дату уже добавлена смена'}), 400

    if start_time != 'Выходной':
        # Проверка 6 смен подряд
        if check_consecutive_shifts(user_id, shift_date) > 6:
            conn.close()
            return jsonify({'error': 'Будет более 6 смен подряд! Максимум — 6.'}), 400

        # Проверка 9 ч отдыха
        ok, msg = check_rest_hours(user_id, shift_date, start_time, end_time)
        if not ok:
            conn.close()
            return jsonify({'error': msg}), 400

    # Сохраняем смену
    conn.execute(
        'INSERT INTO shifts (user_id, shift_date, start_time, end_time) VALUES (?,?,?,?)',
        (user_id, shift_date, start_time, end_time)
    )
    conn.commit()
    shift_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()['id']

    # Предупреждение / запрос на переработку
    overtime_warning = None
    if start_time != 'Выходной' and end_time:
        week_start   = get_week_start(shift_date)
        hours_before = calc_week_hours(user_id, week_start, exclude_id=shift_id)
        new_hours    = shift_duration_hours(start_time, end_time)
        total_hours  = hours_before + new_hours

        if total_hours > WEEK_NORM_HOURS:
            extra = round(total_hours - WEEK_NORM_HOURS, 1)
            # Создаём запрос на переработку
            conn.execute(
                '''INSERT INTO overtime_requests
                   (user_id, shift_id, week_start, planned_hours, extra_hours, created_at)
                   VALUES (?,?,?,?,?,?)''',
                (user_id, shift_id, week_start,
                 round(total_hours, 1), extra,
                 datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            )
            conn.commit()
            overtime_warning = (
                f'Превышение нормы 40 ч/нед: {round(total_hours, 1)} ч '
                f'(+{extra} ч). Запрос отправлен на согласование.'
            )

    conn.close()
    return jsonify({'ok': True, 'overtime_warning': overtime_warning})

@app.route('/shift/delete/<int:shift_id>', methods=['POST'])
@login_required
def delete_shift(shift_id):
    user_id = session['user_id']
    conn = get_db()

    shift = conn.execute('SELECT * FROM shifts WHERE id=?', (shift_id,)).fetchone()
    if not shift:
        conn.close()
        return jsonify({'error': 'Смена не найдена'}), 404
    if session['role'] == 'employee' and shift['user_id'] != user_id:
        conn.close()
        return jsonify({'error': 'Нет доступа'}), 403

    shift_date = shift['shift_date']

    # Удаляем смену
    conn.execute('DELETE FROM shifts WHERE id=?', (shift_id,))
    conn.commit()

    # Уведомление для админов если сотрудник отменяет рабочую смену
    # (уведомление админам на случай если чел умирает от температуры в 37C и отменяет рабочий день)
    if session['role'] == 'employee' and shift['start_time'] != 'Выходной':
        user = conn.execute('SELECT full_name FROM users WHERE id=?', (user_id,)).fetchone()
        name = user['full_name'] if user else session.get('full_name', '')
        create_notification(
            user_id=user_id,
            shift_id=None,  # смена уже удалена
            shift_date=shift_date,
            message=f'Сотрудник {name} отменил рабочую смену {shift_date}.'
        )

    conn.close()
    return jsonify({'ok': True})

@app.route('/shift/edit/<int:shift_id>', methods=['POST'])
@login_required
def edit_shift(shift_id):
    user_id    = session['user_id']
    start_time = request.form['start_time']
    end_time   = request.form.get('end_time', '')

    conn = get_db()
    shift = conn.execute('SELECT * FROM shifts WHERE id=?', (shift_id,)).fetchone()
    if not shift:
        conn.close()
        return jsonify({'error': 'Смена не найдена'}), 404
    if session['role'] == 'employee' and shift['user_id'] != user_id:
        conn.close()
        return jsonify({'error': 'Нет доступа'}), 403

    if start_time != 'Выходной':
        if check_consecutive_shifts(user_id, shift['shift_date'], exclude_id=shift_id) > 6:
            conn.close()
            return jsonify({'error': 'Будет более 6 смен подряд! Максимум — 6.'}), 400

        ok, msg = check_rest_hours(user_id, shift['shift_date'], start_time, end_time, exclude_id=shift_id)
        if not ok:
            conn.close()
            return jsonify({'error': msg}), 400

    conn.execute(
        'UPDATE shifts SET start_time=?, end_time=? WHERE id=?',
        (start_time, end_time, shift_id)
    )
    conn.commit()

    # Пересчёт переработки
    overtime_warning = None
    if start_time != 'Выходной' and end_time:
        week_start   = get_week_start(shift['shift_date'])
        hours_before = calc_week_hours(user_id, week_start, exclude_id=shift_id)
        new_hours    = shift_duration_hours(start_time, end_time)
        total_hours  = hours_before + new_hours
        if total_hours > WEEK_NORM_HOURS:
            extra = round(total_hours - WEEK_NORM_HOURS, 1)
            overtime_warning = (
                f'Превышение нормы 40 ч/нед: {round(total_hours, 1)} ч (+{extra} ч).'
            )

    conn.close()
    return jsonify({'ok': True, 'overtime_warning': overtime_warning})

# ─── Страница руководителя ────────────────────────────────────────────────────

@app.route('/manager')
@login_required
@role_required('admin', 'manager')
def manager_dashboard():
    conn = get_db()

    if session['role'] == 'admin':
        employees = conn.execute(
            "SELECT * FROM users WHERE role='employee' ORDER BY alliance, team, full_name"
        ).fetchall()
    else:
        employees = conn.execute(
            "SELECT * FROM users WHERE role='employee' AND alliance=? ORDER BY team, full_name",
            (session['alliance'],)
        ).fetchall()

    employee_ids = [e['id'] for e in employees]
    shifts_by_user = {}
    if employee_ids:
        ph = ','.join('?' * len(employee_ids))
        for s in conn.execute(
            f'SELECT * FROM shifts WHERE user_id IN ({ph}) ORDER BY shift_date', employee_ids
        ).fetchall():
            shifts_by_user.setdefault(s['user_id'], []).append(dict(s))

    # Уведомления и запросы для панели
    notifications = [dict(r) for r in conn.execute(
        '''SELECT n.*, u.full_name FROM notifications n
           JOIN users u ON u.id=n.user_id
           ORDER BY n.created_at DESC LIMIT 50'''
    ).fetchall()]

    overtime_requests = [dict(r) for r in conn.execute(
        '''SELECT o.*, u.full_name FROM overtime_requests o
           JOIN users u ON u.id=o.user_id
           ORDER BY o.created_at DESC'''
    ).fetchall()]

    unread_count = conn.execute('SELECT COUNT(*) as c FROM notifications WHERE is_read=0').fetchone()['c']

    conn.close()

    employees_data = []
    for emp in employees:
        emp_shifts = shifts_by_user.get(emp['id'], [])
        total = len([s for s in emp_shifts if s['start_time'] != 'Выходной'])
        employees_data.append({**dict(emp), 'shifts': emp_shifts, 'total_shifts': total})

    return render_template('manager.html',
        employees=employees_data,
        user=session,
        dates=get_next_14_days(),
        notifications=notifications,
        overtime_requests=overtime_requests,
        unread_count=unread_count
    )

# ─── API: уведомления и переработки ──────────────────────────────────────────

@app.route('/api/notifications/read', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def mark_notifications_read():
    conn = get_db()
    conn.execute('UPDATE notifications SET is_read=1')
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/overtime/<int:req_id>/review', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def review_overtime(req_id):
    status  = request.json.get('status')   # 'approved' или 'rejected'
    comment = request.json.get('comment', '')
    if status not in ('approved', 'rejected'):
        return jsonify({'error': 'Неверный статус'}), 400
    conn = get_db()
    conn.execute(
        'UPDATE overtime_requests SET status=?, admin_comment=?, reviewed_at=? WHERE id=?',
        (status, comment, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), req_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── API: пользователи ────────────────────────────────────────────────────────

@app.route('/api/employee/<int:emp_id>/shifts')
@login_required
@role_required('admin', 'manager')
def get_employee_shifts(emp_id):
    conn = get_db()
    shifts = [dict(s) for s in conn.execute(
        'SELECT * FROM shifts WHERE user_id=? ORDER BY shift_date', (emp_id,)
    ).fetchall()]
    conn.close()
    return jsonify(shifts)

@app.route('/api/users')
@login_required
@role_required('admin', 'manager')
def get_users():
    conn = get_db()
    if session['role'] == 'admin':
        users = conn.execute(
            'SELECT id,username,full_name,role,alliance,team FROM users ORDER BY role,alliance,full_name'
        ).fetchall()
    else:
        users = conn.execute(
            'SELECT id,username,full_name,role,alliance,team FROM users WHERE alliance=? ORDER BY role,full_name',
            (session['alliance'],)
        ).fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])

@app.route('/api/users/add', methods=['POST'])
@login_required
@role_required('admin', 'manager')
def add_user():
    data = request.json
    password = data.get('password', '')

    # Минимум 8 символов
    if len(password) < MIN_PASSWORD_LEN:
        return jsonify({'error': f'Пароль должен быть не менее {MIN_PASSWORD_LEN} символов'}), 400

    conn = get_db()
    try:
        conn.execute(
            'INSERT INTO users (username,password,full_name,role,alliance,team) VALUES (?,?,?,?,?,?)',
            (data['username'], hash_password(password),
             data['full_name'], data['role'],
             data.get('alliance'), data.get('team'))
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Логин уже занят'}), 400
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/users/delete/<int:user_id>', methods=['POST'])
@login_required
@role_required('admin')
def delete_user(user_id):
    conn = get_db()
    conn.execute('DELETE FROM shifts WHERE user_id=?', (user_id,))
    conn.execute('DELETE FROM notifications WHERE user_id=?', (user_id,))
    conn.execute('DELETE FROM overtime_requests WHERE user_id=?', (user_id,))
    conn.execute('DELETE FROM users WHERE id=?', (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Экспорт Excel ────────────────────────────────────────────────────────────

@app.route('/export/excel')
@login_required
@role_required('admin', 'manager')
def export_excel():
    conn = get_db()
    if session['role'] == 'admin':
        employees = conn.execute(
            "SELECT * FROM users WHERE role='employee' ORDER BY alliance, team, full_name"
        ).fetchall()
    else:
        employees = conn.execute(
            "SELECT * FROM users WHERE role='employee' AND alliance=? ORDER BY team, full_name",
            (session['alliance'],)
        ).fetchall()

    today      = datetime.today()
    date_range = [(today + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(14)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'График смен'

    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    headers = ['Альянс', 'Группа', 'Сотрудник'] + date_range + ['Итого часов']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')

    employee_ids = [e['id'] for e in employees]
    shifts_map = {}
    if employee_ids:
        ph = ','.join('?' * len(employee_ids))
        for s in conn.execute(f'SELECT * FROM shifts WHERE user_id IN ({ph})', employee_ids).fetchall():
            shifts_map[(s['user_id'], s['shift_date'])] = s
    conn.close()

    for emp in employees:
        row = [emp['alliance'] or '', emp['team'] or '', emp['full_name']]
        total_h = 0.0
        for date_str in date_range:
            shift = shifts_map.get((emp['id'], date_str))
            if shift:
                if shift['start_time'] == 'Выходной':
                    row.append('Выходной')
                else:
                    end = shift['end_time'] or ''
                    row.append(f"{shift['start_time']}-{end}" if end else shift['start_time'])
                    total_h += shift_duration_hours(shift['start_time'], end)
            else:
                row.append('')
        row.append(round(total_h, 1))
        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"schedule_{today.strftime('%Y%m%d')}.xlsx"
    )

# ─── Запуск ───────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
